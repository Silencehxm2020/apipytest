from openpyxl import load_workbook

from tools.my_logging import LOG
from tools.get_filepath import GetFilePath


class DOEXL:
    """
    用于表格操作，读取表格头，读取单页表格内容 及读取表格全部内容

    """

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb[self.sheetname]

    # 获取表格第一行
    def get_exl_header(self):
        try:
            LOG.info("读取表格第一行")
            sheet_header = []
            for j in range(1, self.sheet.max_column + 1):
                sheet_header.append(self.sheet.cell(1, j).value)
            return sheet_header
        except Exception as e:
            LOG.error("读取表格第一行失败:{}".format(e))
            raise e

    def get_sheet_data(self):
        """
        读取单页表格内容
        :return: 表格内容以列表套字典方式返回
        """
        try:
            LOG.info("读取表格的{}sheet".format([self.sheetname]))
            case_data = []
            header = self.get_exl_header()
            for i in range(2, self.sheet.max_row + 1):
                case = {}
                for j in range(0, self.sheet.max_column):
                    case[header[j]] = self.sheet.cell(i, j + 1).value
                case_data.append(case)
            return case_data
        except Exception as e:
            LOG.error("读取表格一个sheet失败")
            raise e

    def get_exl_all_data(self):
        """
        读取多sheet页的表格内容，并加上一个字段sheet标识，
        :return: 列表套字典格式输出
        """
        # wb = load_workbook(self.filename)
        try:
            LOG.info("正在读取表格全部数据")
            sheets = self.wb.sheetnames
            case_data = []
            for key in sheets:
                sheet = self.wb[key]
                for i in range(2, sheet.max_row + 1):
                    case = {}
                    for j in range(0, sheet.max_column):
                        case[self.get_exl_header()[j]] = sheet.cell(i, j + 1).value
                    case['sheet'] = key
                    case_data.append(case)
            LOG.info("读取表格数据完成{}".format(case_data))
            return case_data
        except Exception as e:
            LOG.error("读取全部表格数据失败！错误信息为：{}".format(e))
            raise e

    def write_back(self, i, j, value):
        """
        将结果写回Excel，并进行保存
        :param i:行
        :param j:列
        :param value:值
        :return:
        """
        try:
            LOG.info("正在讲数据写回表格")
            self.sheet.cell(i, j).value = value
            self.wb.save(self.filename)
            LOG.info("写回成功")
        except Exception as e:
            LOG.error("数据写入表格失败，失败信息为{}".format(e))

    @staticmethod
    def do_null_and_true(str_dict):
        """
        表格中数据存在null和true使用eval进行字典格式还原时，出现报错。将表格中读出的null和true进行替换
        :return:
        """
        try:
            LOG.info("正在将数据中PYthon无法读取数据替换")
            str_n = str_dict.replace("null", "None")
            str_N = str_n.replace("NULL", "None")
            str_t = str_N.replace("true", "True")
            str_ok = str_t.replace("false", "False")
            LOG.info("替换成功，现数据为：{}".format(str_ok))
            return str_ok
        except Exception as e:
            LOG.error("替换出现错误，错误信息为：{}".format(e))


""

if __name__ == '__main__':
    filepath = GetFilePath.file_path('test_data', 'test_da.xlsx')
    print(filepath)
    data = DOEXL(filepath, 'login').get_sheet_data()
    print(data)
